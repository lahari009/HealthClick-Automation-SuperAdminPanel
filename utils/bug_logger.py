from openpyxl import load_workbook
from datetime import datetime
import os


class BugLogger:

    def __init__(self):
        # Resolve Excel path relative to this file
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_path = os.path.abspath(
            os.path.join(base_dir, "..", "HealthClick_Automation_Bug_Report.xlsx")
        )

    def log_bug(
        self,
        module,
        page,
        scenario,
        steps,
        expected,
        actual,
        severity="High",
        status="FAILED"
    ):
        # Log only failed steps
        if status != "FAILED":
            return

        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(
                f"Bug report Excel not found at: {self.excel_path}"
            )

        workbook = load_workbook(self.excel_path)

        # ✅ Always use FIRST sheet (Bug Report)
        sheet = workbook.worksheets[0]

        bug_id = f"BUG-{sheet.max_row}"

        # ✅ ORDER MATCHES YOUR EXCEL HEADERS EXACTLY
        sheet.append([
            bug_id,                                      # A - Bug ID
            module,                                      # B - Module
            page,                                        # C - Page
            scenario,                                    # D - Scenario
            steps,                                       # E - Steps
            expected,                                    # F - Expected Result
            actual,                                      # G - Actual Result
            severity,                                    # H - Severity
            status,                                      # I - Status
            datetime.now().strftime("%Y-%m-%d %H:%M:%S") # J - Timestamp
        ])

        workbook.save(self.excel_path)
        workbook.close()

        print(f"{bug_id} logged successfully")
